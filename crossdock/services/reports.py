"""Plan efficiency reports (FR-017 savings, FR-018 utilization) + Excel export."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO
from typing import Any

import pandas as pd  # type: ignore[import-untyped]
from sqlalchemy.orm import Session

from crossdock.config import Settings, get_settings
from crossdock.distance.haversine import HaversineDistanceProvider
from crossdock.storage.repositories import AssignmentRepository, OrderRepository, VehicleRepository


@dataclass(frozen=True)
class UtilizationRow:
    vehicle_code: str
    drop_count: int
    distance_km: float
    cost_eur: float
    fill_ratio: float | None
    order_count: int
    total_weight_kg: float
    route_status: str = "proposed"


@dataclass(frozen=True)
class SavingsSummary:
    baseline_cost_eur: float
    optimized_cost_eur: float
    savings_eur: float
    savings_pct: float
    routed_orders: int
    note: str


@dataclass(frozen=True)
class ReportBundle:
    run_id: int
    plan_status: str
    used_fallback_draft: bool
    utilization: tuple[UtilizationRow, ...]
    savings: SavingsSummary
    warnings: tuple[str, ...] = field(default_factory=tuple)
    display_name: str | None = None
    created_at: datetime | None = None


def build_report(
    session: Session,
    *,
    run_id: int | None = None,
    settings: Settings | None = None,
) -> ReportBundle | None:
    """Build utilization + savings for an approved plan (fallback: latest draft)."""
    cfg = settings or get_settings()
    repo = AssignmentRepository(session)
    used_fallback = False
    warnings: list[str] = []

    if run_id is not None:
        run = repo.get_run(run_id)
    else:
        run = repo.get_latest_approved_run()
        if run is None:
            run = repo.get_latest_run()
            if run is not None:
                used_fallback = True
                warnings.append("Brak zatwierdzonego planu — raport z najnowszego draftu.")

    if run is None:
        return None

    routes = repo.list_routes_for_run(run.id)
    items = repo.list_items_for_run(run.id)
    orders = OrderRepository(session)
    vehicles = VehicleRepository(session)
    distance = HaversineDistanceProvider()
    depot = (cfg.depot_latitude, cfg.depot_longitude)

    # Per-vehicle fill from items (prefer item.fill_ratio; else weight/capacity)
    items_by_vehicle: dict[str, list[Any]] = {}
    for item in items:
        if item.vehicle_code in {"UNASSIGNED", "UNROUTED"} or item.sequence is None:
            continue
        items_by_vehicle.setdefault(item.vehicle_code, []).append(item)

    utilization: list[UtilizationRow] = []
    for route in routes:
        v_items = items_by_vehicle.get(route.vehicle_code, [])
        total_weight = sum(i.weight_kg for i in v_items)
        fill: float | None = None
        vehicle = vehicles.get_by_code(route.vehicle_code)
        if vehicle is not None and vehicle.weight_capacity_kg > 0:
            fill = total_weight / vehicle.weight_capacity_kg
            if fill < cfg.min_fill_ratio:
                warnings.append(
                    f"Trasa {route.vehicle_code}: zapełnienie {fill * 100:.0f}% "
                    f"poniżej progu {cfg.min_fill_ratio * 100:.0f}%."
                )
        utilization.append(
            UtilizationRow(
                vehicle_code=route.vehicle_code,
                drop_count=route.drop_count,
                distance_km=route.distance_km,
                cost_eur=route.cost_eur,
                fill_ratio=fill,
                order_count=len(v_items),
                total_weight_kg=total_weight,
                route_status=route.route_status,
            )
        )

    # Baseline: each routed order alone on a round-trip from depot
    baseline = 0.0
    routed_count = 0
    for item in items:
        if item.sequence is None or item.vehicle_code in {"UNASSIGNED", "UNROUTED"}:
            continue
        order = orders.get_by_id(item.order_id)
        if order is None:
            continue
        lat = order.delivery_location.latitude
        lon = order.delivery_location.longitude
        if lat is None or lon is None:
            warnings.append(f"Brak coords dla {item.delivery_code} — pominięto w baseline.")
            continue
        leg = distance.distance_km(depot[0], depot[1], lat, lon)
        baseline += 2.0 * leg * cfg.cost_per_km
        routed_count += 1

    optimized = float(run.total_cost_eur or 0.0)
    savings = baseline - optimized
    savings_pct = (savings / baseline * 100.0) if baseline > 0 else 0.0

    return ReportBundle(
        run_id=run.id,
        plan_status=run.plan_status,
        used_fallback_draft=used_fallback,
        utilization=tuple(utilization),
        savings=SavingsSummary(
            baseline_cost_eur=round(baseline, 2),
            optimized_cost_eur=round(optimized, 2),
            savings_eur=round(savings, 2),
            savings_pct=round(savings_pct, 1),
            routed_orders=routed_count,
            note=(
                "Baseline: 1 zlecenie = 1 pojazd (2x km depot-drop x cost_per_km). "
                "Stawki Sandry (W-06) — placeholder."
            ),
        ),
        warnings=tuple(warnings),
        display_name=run.display_name,
        created_at=run.created_at,
    )


def export_report_xlsx(bundle: ReportBundle) -> bytes:
    """Export report to xlsx bytes (pandas stays inside this function)."""
    from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
    from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

    from crossdock.text_pl import plan_status_pl

    util_rows = [
        {
            "Pojazd": row.vehicle_code,
            "Punkty rozładunku": row.drop_count,
            "Zlecenia": row.order_count,
            "Waga [kg]": round(row.total_weight_kg, 1),
            "Zapełnienie [%]": (
                round(row.fill_ratio * 100, 1) if row.fill_ratio is not None else None
            ),
            "Km": round(row.distance_km, 1),
            "Koszt €": round(row.cost_eur, 2),
        }
        for row in bundle.utilization
    ]
    sav = bundle.savings
    savings_rows = [
        {"Wskaźnik": "Koszt odniesienia €", "Wartość": sav.baseline_cost_eur},
        {"Wskaźnik": "Koszt zoptymalizowany €", "Wartość": sav.optimized_cost_eur},
        {"Wskaźnik": "Oszczędność €", "Wartość": sav.savings_eur},
        {"Wskaźnik": "Oszczędność %", "Wartość": sav.savings_pct},
        {"Wskaźnik": "Zlecenia na trasie", "Wartość": sav.routed_orders},
        {"Wskaźnik": "Numer planu", "Wartość": bundle.run_id},
        {"Wskaźnik": "Status planu", "Wartość": plan_status_pl(bundle.plan_status)},
        {"Wskaźnik": "Uwaga", "Wartość": sav.note},
    ]

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        pd.DataFrame(util_rows).to_excel(writer, sheet_name="Zapełnienie", index=False)
        pd.DataFrame(savings_rows).to_excel(writer, sheet_name="Oszczędności", index=False)
        header_fill = PatternFill("solid", fgColor="0F766E")
        header_font = Font(bold=True, color="FFFFFF")
        for sheet_name in ("Zapełnienie", "Oszczędności"):
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for col_idx, column in enumerate(ws.columns, start=1):
                max_len = 0
                for cell in column:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 48)
            if sheet_name == "Zapełnienie":
                for row in ws.iter_rows(min_row=2, min_col=5, max_col=7):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            if cell.column in (5, 6):
                                cell.number_format = "0.0"
                            else:
                                cell.number_format = "0.00"
    return buffer.getvalue()
